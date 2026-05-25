import re
import urllib.request
import urllib.parse
from datetime import date
from html import unescape

from lxml import html
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = "/Users/yesen/Code/LawLink/outputs/cause_taxonomy/民事刑事行政诉讼案由分级分类表.xlsx"

SOURCES = {
    "civil": {
        "name": "最高人民法院发布修改后的《民事案件案由规定》",
        "authority": "最高人民法院，法〔2025〕226号、法〔2025〕227号，2026年1月1日起施行",
        "url": "https://www.court.gov.cn/zixun/xiangqing/484231.html",
    },
    "admin": {
        "name": "最高人民法院印发《关于行政案件案由的暂行规定》的通知",
        "authority": "最高人民法院，法发〔2020〕44号，2021年1月1日起施行",
        "url": "https://www.court.gov.cn/fabu/xiangqing/282681.html",
    },
    "criminal_base": {
        "name": "中华人民共和国罪名列表（按《刑法》分则章节整理）",
        "authority": "以《中华人民共和国刑法》及两高关于执行刑法确定罪名的规定、补充规定为依据；列表页用于结构化抽取和交叉校验",
        "url": "https://zh.wikipedia.org/wiki/中华人民共和国罪名列表",
    },
    "criminal_supp8": {
        "name": "最高人民法院、最高人民检察院关于执行《中华人民共和国刑法》确定罪名的补充规定（八）",
        "authority": "法释〔2024〕3号，2024年3月1日起施行",
        "url": "https://www.court.gov.cn/zixun/xiangqing/424452.html",
    },
}


def fetch_text(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    raw = urllib.request.urlopen(req, timeout=30).read()
    doc = html.fromstring(raw)
    for bad in doc.xpath("//script|//style|//noscript"):
        bad.drop_tree()
    lines = []
    for text in doc.text_content().splitlines():
        text = re.sub(r"\s+", " ", unescape(text)).strip()
        if text:
            lines.append(text)
    return lines


def strip_num(text, pattern):
    m = re.match(pattern, text)
    return (m.group(1), m.group(2).strip()) if m else ("", text.strip())


def parse_civil():
    lines = fetch_text(SOURCES["civil"]["url"])
    start = next(i for i, line in enumerate(lines) if line.startswith("为了正确适用法律，统一确定案由"))
    lines = lines[start + 1 :]
    rows = []
    first_no = first = second_no = second = third_no = third = ""
    for line in lines:
        if line.startswith("责任编辑") or "版权所有" in line:
            break
        if re.match(r"^第[一二三四五六七八九十]+部分\s+", line):
            first_no, first = strip_num(line, r"^(第[一二三四五六七八九十]+部分)\s+(.+)$")
            second_no = second = third_no = third = ""
            rows.append(["民事", 1, first_no, first, "", "", "", "", "", ""])
        elif re.match(r"^[一二三四五六七八九十]+、", line):
            second_no, second = strip_num(line, r"^([一二三四五六七八九十]+)、(.+)$")
            third_no = third = ""
            rows.append(["民事", 2, first_no, first, second_no, second, "", "", "", ""])
        elif re.match(r"^\d+\.", line):
            third_no, third = strip_num(line, r"^(\d+)\.(.+)$")
            rows.append(["民事", 3, first_no, first, second_no, second, third_no, third, "", ""])
        elif re.match(r"^（\d+）", line):
            fourth_no, fourth = strip_num(line, r"^(（\d+）)(.+)$")
            rows.append(["民事", 4, first_no, first, second_no, second, third_no, third, fourth_no, fourth])
    return rows


def parse_admin():
    lines = fetch_text(SOURCES["admin"]["url"])
    start = next(i for i, line in enumerate(lines) if line == "一级案由")
    rows = []
    first = "行政行为"
    second_no = second = ""
    rows.append(["行政", 1, "", first, "", "", ""])
    for line in lines[start + 1 :]:
        if line.startswith("责任编辑"):
            break
        if line in {"行政行为", "二级、三级案由"}:
            continue
        if re.match(r"^[（(][一二三四五六七八九十]+[）)]", line):
            second_no, second = strip_num(line, r"^([（(][一二三四五六七八九十]+[）)])(.+)$")
            rows.append(["行政", 2, "", first, second_no, second, ""])
        elif re.match(r"^\d+\.", line):
            third_no, third = strip_num(line, r"^(\d+)\.(.+)$")
            rows.append(["行政", 3, "", first, second_no, second, third_no + "." + third])
    special_rules = [
        ["行政", "特殊规则", "行政复议案件", "复议维持或实体驳回复议申请的，表述为“××（行政行为）及行政复议”。"],
        ["行政", "特殊规则", "行政协议案件", "须列明行政协议名称；赔偿、解除、继续履行等请求可一并列出。"],
        ["行政", "特殊规则", "行政赔偿案件", "一并提起的表述为“××（行政行为）及行政赔偿”；单独提起的表述为“行政赔偿”。"],
        ["行政", "特殊规则", "规范性文件审查", "表述为“××（行政行为）及规范性文件审查”。"],
        ["行政", "特殊规则", "行政公益诉讼", "表述为“××（行政行为）公益诉讼”。"],
        ["行政", "特殊规则", "不履行法定职责", "表述为“不履行××职责”。"],
        ["行政", "特殊规则", "申请执行生效法律文书", "表述为“申请执行”+行政诉讼案由+“判决/裁定/调解书”。"],
        ["行政", "特殊规则", "非诉行政执行", "表述为“申请执行××（行政行为）”。"],
    ]
    return rows, special_rules


def parse_criminal():
    url = urllib.parse.quote(SOURCES["criminal_base"]["url"], safe=":/?#[]@!$&'()*+,;=%")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    doc = html.fromstring(urllib.request.urlopen(req, timeout=30).read())
    content = doc.xpath("//div[contains(@class,'mw-parser-output')]")[0]
    chapter_names = {
        "危害国家安全罪",
        "危害公共安全罪",
        "破坏社会主义市场经济秩序罪",
        "侵犯公民人身权利、民主权利罪",
        "侵犯财产罪",
        "妨害社会管理秩序罪",
        "危害国防利益罪",
        "贪污贿赂罪",
        "渎职罪",
        "军人违反职责罪",
    }
    rows = []
    chapter = section = ""
    seq = 0
    missing_refs = {
        "违规披露、不披露重要信息罪": "第161条",
        "非国家工作人员受贿罪": "第163条",
        "对非国家工作人员行贿罪": "第164条第1款",
        "组织、利用会道门、邪教组织、利用迷信致人重伤、死亡罪": "第300条第2款",
        "拒绝提供间谍犯罪、恐怖主义犯罪、极端主义犯罪证据罪": "第311条",
    }
    for el in content.xpath(".//*[self::h2 or self::h3 or self::ul]"):
        tag = el.tag.lower() if isinstance(el.tag, str) else ""
        if not tag:
            continue
        text = re.sub(r"\[编辑\]", "", el.text_content()).strip()
        text = re.sub(r"\s+", " ", text)
        if tag == "h2":
            heading = re.sub(r"^\d+\s*", "", text)
            if heading in chapter_names:
                chapter, section = heading, ""
            elif heading == "參考文獻":
                break
        elif tag == "h3" and chapter:
            section = re.sub(r"^\d+(\.\d+)?\s*", "", text)
        elif tag == "ul" and chapter:
            for li in el.xpath("./li"):
                item = re.sub(r"\[[^\]]+\]", "", li.text_content())
                item = re.sub(r"\s+", " ", item).strip()
                if not item or "已取消" in item or "已替换" in item:
                    continue
                m = re.match(r"(.+?)（([^（）]+)）", item)
                if m:
                    seq += 1
                    rows.append(["刑事", seq, chapter, section, m.group(1).strip(), m.group(2).strip()])
                elif item in missing_refs:
                    seq += 1
                    rows.append(["刑事", seq, chapter, section, item, missing_refs[item]])
    return rows


def add_sheet_table(ws, headers, rows, table_name):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    end_col = get_column_letter(len(headers))
    end_row = max(ws.max_row, 1)
    tab = Table(displayName=table_name, ref=f"A1:{end_col}{end_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{end_col}{end_row}"


def style_workbook(path):
    wb = load_workbook(path)
    dark_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = dark_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.sheet_view.showGridLines = False
        widths = {}
        for row in ws.iter_rows(values_only=True):
            for idx, value in enumerate(row, 1):
                value = "" if value is None else str(value)
                widths[idx] = min(max(widths.get(idx, 8), len(value) + 2), 48)
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.row_dimensions[1].height = 28
    wb.save(path)


def main():
    civil_rows = parse_civil()
    admin_rows, admin_rules = parse_admin()
    criminal_rows = parse_criminal()

    wb = Workbook()
    ws = wb.active
    ws.title = "编制说明"
    ws.append(["项目", "内容"])
    notes = [
        ("编制日期", date.today().isoformat()),
        ("总体口径", "民事、行政按最高人民法院案由规定整理；刑事诉讼实务案由通常体现为罪名，按《刑法》分则章/节/罪名整理。"),
        ("民事案由口径", "采用最高人民法院 2025 年第三次修正后的《民事案件案由规定》，保留第一级至第四级。"),
        ("刑事案由口径", "采用刑法罪名体系，保留刑法分则章、节、罪名、对应条文。已取消或已替换罪名未列入。"),
        ("行政案由口径", "采用最高人民法院《关于行政案件案由的暂行规定》，保留一级、二级、三级，并另列特殊确定规则。"),
        ("民事案由数量", str(len(civil_rows))),
        ("刑事罪名数量", str(len(criminal_rows))),
        ("行政案由条目数量", str(len(admin_rows))),
    ]
    for row in notes:
        ws.append(row)

    add_sheet_table(
        wb.create_sheet("民事案由"),
        ["诉讼类型", "层级", "一级编号", "一级案由", "二级编号", "二级案由", "三级编号", "三级案由", "四级编号", "四级案由"],
        civil_rows,
        "CivilCauses",
    )
    add_sheet_table(
        wb.create_sheet("刑事罪名"),
        ["诉讼类型", "序号", "刑法分则章", "刑法分则节", "罪名", "对应条文"],
        criminal_rows,
        "CriminalCharges",
    )
    add_sheet_table(
        wb.create_sheet("行政案由"),
        ["诉讼类型", "层级", "一级编号", "一级案由", "二级编号", "二级案由", "三级案由"],
        admin_rows,
        "AdministrativeCauses",
    )
    add_sheet_table(
        wb.create_sheet("行政特殊规则"),
        ["诉讼类型", "类型", "规则名称", "案由确定规则"],
        admin_rules,
        "AdministrativeRules",
    )
    source_rows = [[k, v["name"], v["authority"], v["url"]] for k, v in SOURCES.items()]
    add_sheet_table(
        wb.create_sheet("来源索引"),
        ["来源键", "文件/页面", "效力与口径", "URL"],
        source_rows,
        "SourceIndex",
    )
    ws.freeze_panes = "A2"
    wb.save(OUTPUT)
    style_workbook(OUTPUT)
    print(OUTPUT)
    print(f"civil={len(civil_rows)} criminal={len(criminal_rows)} admin={len(admin_rows)}")


if __name__ == "__main__":
    main()
